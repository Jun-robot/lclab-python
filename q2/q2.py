import openpyxl as px

def create_excel():
  wb  = px.Workbook()
  
  ws1 = wb.active
  ws1.title = 'q2_sheet1'

  ws2 = wb.create_sheet(title='q2_sheet2')

  #q2_sheet1: 1行目に1~10を入力
  for col in range(1, 11):
    ws1.cell(row=1, column=col, value=col)

  #q2_sheet2: 1列目に等比数列の値を入力
  for row in range(1, 11):
    ws2.cell(row=row, column=1, value=2*(5**(row-1)))
  
  wb.save('q2.xlsx')

def main():
  create_excel()

if __name__ == '__main__':
  main()